"""
Excel tracker for project-note-tracker plugin.

Manages a tracker.xlsx file with two sheets:
  Questions: Handler | Question | Internal Review | Handler Answer | Status | Date Added
  Bugs:      Summary | Severity | Steps to Reproduce | Investigation | Status | Date Reported

Usage:
  tracker.py init <dir> [handler ...]
  tracker.py add <dir> <handler> <question> <internal_review> <status>
  tracker.py pending <dir> [--handler <name>]
  tracker.py resolve <dir> <row> <answer>
  tracker.py decide <dir> <row> <decision+rationale>
  tracker.py add-handler <dir> <handler>
  tracker.py list-handlers <dir>
  tracker.py update-review <dir> <row> <internal_review> <status>
  tracker.py doctor <dir>
  tracker.py add-bug <dir> <summary> <severity> <steps> <investigation> <status>
  tracker.py list-bugs <dir> [--status <s>] [--severity <s>] [--all]
  tracker.py update-bug <dir> <row> <investigation> <status>
  tracker.py resolve-bug <dir> <row> <resolution>

Requires: openpyxl (run via uvx --with openpyxl)
"""

import json
import sys
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.styles.differential import DifferentialStyle
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    print("openpyxl not available. Run via: uvx --with openpyxl python3 tracker.py ...", file=sys.stderr)
    sys.exit(1)

# ---------------------------------------------------------------------------
# Questions sheet constants
# ---------------------------------------------------------------------------
COLUMNS = ["Handler", "Question", "Internal Review", "Handler Answer", "Status", "Date Added"]
STATUS_ANSWERED = "Answered Internally"
STATUS_PENDING = "Pending"
STATUS_COMPLETED = "Completed"
STATUS_DECIDED = "Decided"
STATUS_LIST = f'"{STATUS_ANSWERED},{STATUS_PENDING},{STATUS_COMPLETED},{STATUS_DECIDED}"'
Q_HANDLER_COL = 1     # Column A: Handler
Q_QUESTION_COL = 2    # Column B: Question
Q_REVIEW_COL = 3      # Column C: Internal Review
Q_ANSWER_COL = 4      # Column D: Handler Answer
STATUS_COL_INDEX = 5  # column E
STATUS_COL_RANGE = "E2:E1000"

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT_WHITE = Font(bold=True, size=11, color="FFFFFF")
WRAP_ALIGNMENT = Alignment(wrap_text=True, vertical="top")
COL_WIDTHS = {"A": 18, "B": 40, "C": 50, "D": 50, "E": 22, "F": 14}

# Conditional formatting — uses DifferentialStyle so Excel evaluates live
CF_RULES = [
    (STATUS_ANSWERED, DifferentialStyle(
        font=Font(color="375623", bold=True),
        fill=PatternFill(bgColor="E2EFDA"),
    )),
    (STATUS_PENDING, DifferentialStyle(
        font=Font(color="BF8F00", bold=True),
        fill=PatternFill(bgColor="FCE4D6"),
    )),
    (STATUS_COMPLETED, DifferentialStyle(
        font=Font(color="2F5496", bold=True),
        fill=PatternFill(bgColor="D9E2F3"),
    )),
    (STATUS_DECIDED, DifferentialStyle(
        font=Font(color="7030A0", bold=True),
        fill=PatternFill(bgColor="E8D5F5"),
    )),
]

# ---------------------------------------------------------------------------
# Bugs sheet constants
# ---------------------------------------------------------------------------
BUG_COLUMNS = ["Summary", "Severity", "Steps to Reproduce", "Investigation", "Status", "Date Reported"]
BUG_STATUS_OPEN = "Open"
BUG_STATUS_INVESTIGATING = "Investigating"
BUG_STATUS_REPRODUCED = "Reproduced"
BUG_STATUS_FIXED = "Fixed"
BUG_STATUS_CLOSED = "Closed"
BUG_STATUS_LIST = '"Open,Investigating,Reproduced,Fixed,Closed"'
BUG_SEVERITY_LIST = '"Critical,High,Medium,Low"'
BUG_SUMMARY_COL = 1        # Column A: Bug Summary
BUG_SEVERITY_COL_INDEX = 2  # column B
BUG_DESCRIPTION_COL = 3    # Column C: Description
BUG_INVESTIGATION_COL = 4  # Column D: Investigation
BUG_STATUS_COL_INDEX = 5   # column E
BUG_STATUS_COL_RANGE = "E2:E1000"
BUG_SEVERITY_COL_RANGE = "B2:B1000"

BUG_HEADER_FILL = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
BUG_COL_WIDTHS = {"A": 50, "B": 14, "C": 40, "D": 50, "E": 18, "F": 14}

BUG_CF_STATUS = [
    (BUG_STATUS_OPEN, DifferentialStyle(
        font=Font(color="CC0000", bold=True),
        fill=PatternFill(bgColor="FFCCCC"),
    )),
    (BUG_STATUS_INVESTIGATING, DifferentialStyle(
        font=Font(color="BF8F00", bold=True),
        fill=PatternFill(bgColor="FCE4D6"),
    )),
    (BUG_STATUS_REPRODUCED, DifferentialStyle(
        font=Font(color="375623", bold=True),
        fill=PatternFill(bgColor="E2EFDA"),
    )),
    (BUG_STATUS_FIXED, DifferentialStyle(
        font=Font(color="2F5496", bold=True),
        fill=PatternFill(bgColor="D9E2F3"),
    )),
    (BUG_STATUS_CLOSED, DifferentialStyle(
        font=Font(color="808080", bold=True),
        fill=PatternFill(bgColor="D9D9D9"),
    )),
]

BUG_CF_SEVERITY = [
    ("Critical", DifferentialStyle(
        font=Font(color="CC0000", bold=True),
        fill=PatternFill(bgColor="FFCCCC"),
    )),
    ("High", DifferentialStyle(
        font=Font(color="BF8F00", bold=True),
        fill=PatternFill(bgColor="FCE4D6"),
    )),
    ("Medium", DifferentialStyle(
        font=Font(color="375623", bold=True),
        fill=PatternFill(bgColor="E2EFDA"),
    )),
    ("Low", DifferentialStyle(
        font=Font(color="2F5496", bold=True),
        fill=PatternFill(bgColor="D9E2F3"),
    )),
]


# ---------------------------------------------------------------------------
# Common helpers
# ---------------------------------------------------------------------------
def _tracker_path(directory: str) -> Path:
    return Path(directory) / "tracker.xlsx"


# ---------------------------------------------------------------------------
# Questions sheet helpers
# ---------------------------------------------------------------------------
def _apply_conditional_formatting(ws) -> None:
    """Set up conditional formatting rules on the Status column.

    These are real Excel conditional formatting rules — colors update
    automatically when the user changes the dropdown value in Excel.
    """
    from collections import OrderedDict

    from openpyxl.formatting.rule import Rule

    # Clear ALL existing conditional formatting to avoid duplicates on doctor re-runs
    ws.conditional_formatting._cf_rules = OrderedDict()

    for status_value, dxf in CF_RULES:
        rule = Rule(type="cellIs", operator="equal", dxf=dxf, formula=[f'"{status_value}"'])
        ws.conditional_formatting.add(STATUS_COL_RANGE, rule)


def _add_status_dropdown(ws, row: int) -> None:
    """Add data validation dropdown for the status cell in a given row."""
    dv = DataValidation(type="list", formula1=STATUS_LIST, allow_blank=False)
    dv.error = "Pick a valid status"
    dv.errorTitle = "Invalid Status"
    dv.prompt = "Select status"
    dv.promptTitle = "Status"
    dv.add(f"E{row}")
    ws.add_data_validation(dv)


def _add_status_dropdown_range(ws) -> None:
    """Add a single data validation dropdown covering the entire status column."""
    # Clear existing validations on column E
    ws.data_validations.dataValidation = [
        dv for dv in ws.data_validations.dataValidation
        if not any("E" in str(s) for s in (dv.sqref.ranges if hasattr(dv.sqref, 'ranges') else [dv.sqref]))
    ]

    dv = DataValidation(type="list", formula1=STATUS_LIST, allow_blank=False)
    dv.error = "Pick a valid status"
    dv.errorTitle = "Invalid Status"
    dv.prompt = "Select status"
    dv.promptTitle = "Status"
    dv.add(STATUS_COL_RANGE)
    ws.add_data_validation(dv)


def _style_headers(ws) -> None:
    for col_idx, col_name in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = HEADER_FONT_WHITE
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for col_letter, width in COL_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width


def _setup_sheet(ws) -> None:
    """Apply all sheet-level formatting: headers, conditional formatting, dropdown, filter, freeze."""
    _style_headers(ws)
    _apply_conditional_formatting(ws)
    _add_status_dropdown_range(ws)
    ws.auto_filter.ref = "A1:F1"
    ws.freeze_panes = "A2"


# ---------------------------------------------------------------------------
# Bugs sheet helpers
# ---------------------------------------------------------------------------
def _apply_bug_conditional_formatting(ws) -> None:
    """Set up conditional formatting on Status and Severity columns for the Bugs sheet."""
    from collections import OrderedDict

    from openpyxl.formatting.rule import Rule

    ws.conditional_formatting._cf_rules = OrderedDict()

    # Status column (E)
    for status_value, dxf in BUG_CF_STATUS:
        rule = Rule(type="cellIs", operator="equal", dxf=dxf, formula=[f'"{status_value}"'])
        ws.conditional_formatting.add(BUG_STATUS_COL_RANGE, rule)

    # Severity column (B)
    for severity_value, dxf in BUG_CF_SEVERITY:
        rule = Rule(type="cellIs", operator="equal", dxf=dxf, formula=[f'"{severity_value}"'])
        ws.conditional_formatting.add(BUG_SEVERITY_COL_RANGE, rule)


def _add_bug_dropdowns(ws) -> None:
    """Add dropdowns for Status and Severity columns on the Bugs sheet."""
    ws.data_validations.dataValidation = []

    # Status dropdown (column E)
    dv_status = DataValidation(type="list", formula1=BUG_STATUS_LIST, allow_blank=False)
    dv_status.error = "Pick a valid status"
    dv_status.errorTitle = "Invalid Status"
    dv_status.prompt = "Select status"
    dv_status.promptTitle = "Bug Status"
    dv_status.add(BUG_STATUS_COL_RANGE)
    ws.add_data_validation(dv_status)

    # Severity dropdown (column B)
    dv_severity = DataValidation(type="list", formula1=BUG_SEVERITY_LIST, allow_blank=False)
    dv_severity.error = "Pick a valid severity"
    dv_severity.errorTitle = "Invalid Severity"
    dv_severity.prompt = "Select severity"
    dv_severity.promptTitle = "Bug Severity"
    dv_severity.add(BUG_SEVERITY_COL_RANGE)
    ws.add_data_validation(dv_severity)


def _style_bug_headers(ws) -> None:
    """Style the Bugs sheet headers with dark red fill."""
    for col_idx, col_name in enumerate(BUG_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = HEADER_FONT_WHITE
        cell.fill = BUG_HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for col_letter, width in BUG_COL_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width


def _setup_bugs_sheet(ws) -> None:
    """Apply all formatting to the Bugs sheet."""
    _style_bug_headers(ws)
    _apply_bug_conditional_formatting(ws)
    _add_bug_dropdowns(ws)
    ws.auto_filter.ref = "A1:F1"
    ws.freeze_panes = "A2"


def _get_or_create_bugs_sheet(wb):
    """Get the Bugs sheet, creating and formatting it if it doesn't exist."""
    if "Bugs" in wb.sheetnames:
        return wb["Bugs"]
    ws = wb.create_sheet("Bugs")
    _setup_bugs_sheet(ws)
    return ws


# ---------------------------------------------------------------------------
# Questions commands
# ---------------------------------------------------------------------------
def cmd_init(directory: str, handlers: list[str]) -> None:
    """Initialize project-notes directory with tracker.xlsx and handler folders."""
    base = Path(directory)
    base.mkdir(parents=True, exist_ok=True)

    tracker = _tracker_path(directory)
    if tracker.exists():
        print(f"tracker.xlsx already exists at {tracker}", file=sys.stderr)
        sys.exit(1)

    wb = Workbook()
    ws = wb.active
    ws.title = "Questions"
    _setup_sheet(ws)
    wb.save(tracker)

    for handler in handlers:
        handler_dir = base / handler.lower().replace(" ", "-")
        handler_dir.mkdir(exist_ok=True)
        research_file = handler_dir / "research.md"
        if not research_file.exists():
            research_file.write_text(
                f"## Research instructions for {handler} questions\n\n"
                "<!-- Describe where Claude should look and what matters for this handler -->\n"
                "<!-- Examples: -->\n"
                "<!-- - Look in Technical-docs/ for process flows -->\n"
                "<!-- - Check .scout-index.json files for schema context -->\n"
                "<!-- - This team cares about: SLAs, timelines, costs -->\n"
                "<!-- - Terminology: they say X not Y -->\n",
                encoding="utf-8",
            )

    print(json.dumps({"status": "ok", "tracker": str(tracker), "handlers": handlers}))


def cmd_add(directory: str, handler: str, question: str, internal_review: str, status: str) -> None:
    """Append a question row to tracker.xlsx."""
    tracker = _tracker_path(directory)
    if not tracker.exists():
        print("tracker.xlsx not found. Run init first.", file=sys.stderr)
        sys.exit(1)

    wb = load_workbook(tracker)
    ws = wb.active
    row = [handler, question, internal_review, "", status, datetime.now().strftime("%Y-%m-%d")]
    ws.append(row)
    row_num = ws.max_row
    for col_idx in range(1, len(COLUMNS) + 1):
        ws.cell(row=row_num, column=col_idx).alignment = WRAP_ALIGNMENT
    ws.cell(row=row_num, column=STATUS_COL_INDEX).alignment = Alignment(horizontal="center", vertical="top")
    wb.save(tracker)

    print(json.dumps({"status": "ok", "row": row_num, "handler": handler, "question": question}))


def cmd_pending(directory: str, handler_filter: str | None = None) -> None:
    """List pending/unanswered questions as JSON."""
    tracker = _tracker_path(directory)
    if not tracker.exists():
        print("tracker.xlsx not found.", file=sys.stderr)
        sys.exit(1)

    wb = load_workbook(tracker, read_only=True)
    ws = wb.active
    results = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
        values = [cell.value for cell in row]
        if len(values) < 5:
            continue
        handler, question, review, answer, status = values[0], values[1], values[2], values[3], values[4]
        if status and status.strip() in (STATUS_COMPLETED, STATUS_DECIDED):
            continue
        if handler_filter and handler and handler.lower() != handler_filter.lower():
            continue
        results.append({
            "row": row_idx,
            "handler": handler or "",
            "question": question or "",
            "internal_review": review or "",
            "handler_answer": answer or "",
            "status": status or "",
            "date": values[5] if len(values) > 5 else "",
        })
    wb.close()

    print(json.dumps(results, indent=2))


def cmd_resolve(directory: str, row_num: int, answer: str) -> None:
    """Set a row's Handler Answer and mark as Completed."""
    tracker = _tracker_path(directory)
    if not tracker.exists():
        print("tracker.xlsx not found.", file=sys.stderr)
        sys.exit(1)

    wb = load_workbook(tracker)
    ws = wb.active
    if row_num < 2 or row_num > ws.max_row:
        print(f"Row {row_num} out of range (2-{ws.max_row}).", file=sys.stderr)
        sys.exit(1)

    ws.cell(row=row_num, column=Q_ANSWER_COL, value=answer).alignment = WRAP_ALIGNMENT
    ws.cell(row=row_num, column=STATUS_COL_INDEX, value=STATUS_COMPLETED).alignment = Alignment(
        horizontal="center", vertical="top"
    )
    wb.save(tracker)

    question = ws.cell(row=row_num, column=Q_QUESTION_COL).value
    print(json.dumps({"status": "ok", "row": row_num, "question": question, "marked": STATUS_COMPLETED}))


def cmd_decide(directory: str, row_num: int, decision: str) -> None:
    """Set a row's Handler Answer to the decision + rationale and mark as Decided."""
    tracker = _tracker_path(directory)
    if not tracker.exists():
        print("tracker.xlsx not found.", file=sys.stderr)
        sys.exit(1)

    wb = load_workbook(tracker)
    ws = wb.active
    if row_num < 2 or row_num > ws.max_row:
        print(f"Row {row_num} out of range (2-{ws.max_row}).", file=sys.stderr)
        sys.exit(1)

    ws.cell(row=row_num, column=Q_ANSWER_COL, value=decision).alignment = WRAP_ALIGNMENT
    ws.cell(row=row_num, column=STATUS_COL_INDEX, value=STATUS_DECIDED).alignment = Alignment(
        horizontal="center", vertical="top"
    )
    wb.save(tracker)

    question = ws.cell(row=row_num, column=Q_QUESTION_COL).value
    print(json.dumps({"status": "ok", "row": row_num, "question": question, "marked": STATUS_DECIDED}))


def cmd_update_review(directory: str, row_num: int, internal_review: str, status: str) -> None:
    """Update the Internal Review and Status of an existing row."""
    tracker = _tracker_path(directory)
    if not tracker.exists():
        print("tracker.xlsx not found.", file=sys.stderr)
        sys.exit(1)

    wb = load_workbook(tracker)
    ws = wb.active
    if row_num < 2 or row_num > ws.max_row:
        print(f"Row {row_num} out of range (2-{ws.max_row}).", file=sys.stderr)
        sys.exit(1)

    ws.cell(row=row_num, column=Q_REVIEW_COL, value=internal_review).alignment = WRAP_ALIGNMENT
    ws.cell(row=row_num, column=STATUS_COL_INDEX, value=status).alignment = Alignment(
        horizontal="center", vertical="top"
    )
    wb.save(tracker)

    question = ws.cell(row=row_num, column=Q_QUESTION_COL).value
    print(json.dumps({"status": "ok", "row": row_num, "question": question, "new_status": status}))


def cmd_doctor(directory: str) -> None:
    """Upgrade existing tracker.xlsx to latest formatting."""
    tracker = _tracker_path(directory)
    if not tracker.exists():
        print("tracker.xlsx not found.", file=sys.stderr)
        sys.exit(1)

    wb = load_workbook(tracker)
    ws = wb.active

    # Apply all sheet-level formatting (headers, conditional formatting, dropdown, filter, freeze)
    _setup_sheet(ws)

    fixes = 0
    for row in range(2, ws.max_row + 1):
        status_cell = ws.cell(row=row, column=STATUS_COL_INDEX)
        if not status_cell.value:
            continue

        # Center-align the status cell
        status_cell.alignment = Alignment(horizontal="center", vertical="top")

        # Apply wrap alignment to all cells in the row
        for col_idx in range(1, len(COLUMNS) + 1):
            if col_idx != STATUS_COL_INDEX:
                ws.cell(row=row, column=col_idx).alignment = WRAP_ALIGNMENT

        fixes += 1

    # Also doctor the Bugs sheet if it exists
    bug_fixes = 0
    if "Bugs" in wb.sheetnames:
        bugs_ws = wb["Bugs"]
        _setup_bugs_sheet(bugs_ws)
        for row in range(2, bugs_ws.max_row + 1):
            status_cell = bugs_ws.cell(row=row, column=BUG_STATUS_COL_INDEX)
            if not status_cell.value:
                continue
            status_cell.alignment = Alignment(horizontal="center", vertical="top")
            severity_cell = bugs_ws.cell(row=row, column=BUG_SEVERITY_COL_INDEX)
            severity_cell.alignment = Alignment(horizontal="center", vertical="top")
            for col_idx in range(1, len(BUG_COLUMNS) + 1):
                if col_idx not in (BUG_STATUS_COL_INDEX, BUG_SEVERITY_COL_INDEX):
                    bugs_ws.cell(row=row, column=col_idx).alignment = WRAP_ALIGNMENT
            bug_fixes += 1

    wb.save(tracker)
    print(json.dumps({
        "status": "ok",
        "rows_updated": fixes,
        "bug_rows_updated": bug_fixes,
        "tracker": str(tracker),
    }))


def cmd_add_handler(directory: str, handler: str) -> None:
    """Create a new handler directory with research.md template."""
    base = Path(directory)
    if not base.exists():
        print(f"Directory {directory} not found. Run init first.", file=sys.stderr)
        sys.exit(1)

    handler_dir = base / handler.lower().replace(" ", "-")
    handler_dir.mkdir(exist_ok=True)
    research_file = handler_dir / "research.md"
    if not research_file.exists():
        research_file.write_text(
            f"## Research instructions for {handler} questions\n\n"
            "<!-- Describe where Claude should look and what matters for this handler -->\n",
            encoding="utf-8",
        )

    print(json.dumps({"status": "ok", "handler": handler, "path": str(handler_dir)}))


def cmd_list_handlers(directory: str) -> None:
    """List all handler directories."""
    base = Path(directory)
    if not base.exists():
        print("[]")
        return

    handlers = []
    for d in sorted(base.iterdir()):
        if d.is_dir() and not d.name.startswith(".") and not d.name.startswith("_"):
            has_research = (d / "research.md").exists()
            handlers.append({"name": d.name, "has_research": has_research})

    print(json.dumps(handlers, indent=2))


# ---------------------------------------------------------------------------
# Bug commands
# ---------------------------------------------------------------------------
def cmd_add_bug(directory: str, summary: str, severity: str, steps: str,
                investigation: str, status: str) -> None:
    """Append a bug row to the Bugs sheet in tracker.xlsx."""
    tracker = _tracker_path(directory)
    if not tracker.exists():
        print("tracker.xlsx not found. Run init first.", file=sys.stderr)
        sys.exit(1)

    wb = load_workbook(tracker)
    ws = _get_or_create_bugs_sheet(wb)
    row = [summary, severity, steps, investigation, status, datetime.now().strftime("%Y-%m-%d")]
    ws.append(row)
    row_num = ws.max_row
    for col_idx in range(1, len(BUG_COLUMNS) + 1):
        ws.cell(row=row_num, column=col_idx).alignment = WRAP_ALIGNMENT
    ws.cell(row=row_num, column=BUG_STATUS_COL_INDEX).alignment = Alignment(
        horizontal="center", vertical="top"
    )
    ws.cell(row=row_num, column=BUG_SEVERITY_COL_INDEX).alignment = Alignment(
        horizontal="center", vertical="top"
    )
    wb.save(tracker)

    print(json.dumps({"status": "ok", "row": row_num, "summary": summary, "severity": severity}))


def cmd_list_bugs(directory: str, status_filter: str | None = None,
                  severity_filter: str | None = None, show_all: bool = False) -> None:
    """List bugs from the Bugs sheet as JSON."""
    tracker = _tracker_path(directory)
    if not tracker.exists():
        print("tracker.xlsx not found.", file=sys.stderr)
        sys.exit(1)

    wb = load_workbook(tracker, read_only=True)
    if "Bugs" not in wb.sheetnames:
        wb.close()
        print("[]")
        return

    ws = wb["Bugs"]
    results = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
        values = [cell.value for cell in row]
        if len(values) < 5:
            continue
        summary, severity, steps, investigation, status = (
            values[0], values[1], values[2], values[3], values[4],
        )
        if not summary:
            continue
        # By default, hide Closed bugs unless --all
        if not show_all and status and status.strip() == BUG_STATUS_CLOSED:
            continue
        if status_filter and status and status.strip().lower() != status_filter.lower():
            continue
        if severity_filter and severity and severity.strip().lower() != severity_filter.lower():
            continue
        results.append({
            "row": row_idx,
            "summary": summary or "",
            "severity": severity or "",
            "steps": steps or "",
            "investigation": investigation or "",
            "status": status or "",
            "date": values[5] if len(values) > 5 else "",
        })
    wb.close()

    print(json.dumps(results, indent=2))


def cmd_update_bug(directory: str, row_num: int, investigation: str, status: str) -> None:
    """Update the Investigation and Status of an existing bug row."""
    tracker = _tracker_path(directory)
    if not tracker.exists():
        print("tracker.xlsx not found.", file=sys.stderr)
        sys.exit(1)

    wb = load_workbook(tracker)
    if "Bugs" not in wb.sheetnames:
        print("No Bugs sheet found in tracker.xlsx.", file=sys.stderr)
        sys.exit(1)

    ws = wb["Bugs"]
    if row_num < 2 or row_num > ws.max_row:
        print(f"Row {row_num} out of range (2-{ws.max_row}).", file=sys.stderr)
        sys.exit(1)

    ws.cell(row=row_num, column=BUG_INVESTIGATION_COL, value=investigation).alignment = WRAP_ALIGNMENT
    ws.cell(row=row_num, column=BUG_STATUS_COL_INDEX, value=status).alignment = Alignment(
        horizontal="center", vertical="top"
    )
    wb.save(tracker)

    summary = ws.cell(row=row_num, column=BUG_SUMMARY_COL).value
    print(json.dumps({"status": "ok", "row": row_num, "summary": summary, "new_status": status}))


def cmd_resolve_bug(directory: str, row_num: int, resolution: str) -> None:
    """Mark a bug as Fixed with a resolution note in the Investigation column."""
    tracker = _tracker_path(directory)
    if not tracker.exists():
        print("tracker.xlsx not found.", file=sys.stderr)
        sys.exit(1)

    wb = load_workbook(tracker)
    if "Bugs" not in wb.sheetnames:
        print("No Bugs sheet found in tracker.xlsx.", file=sys.stderr)
        sys.exit(1)

    ws = wb["Bugs"]
    if row_num < 2 or row_num > ws.max_row:
        print(f"Row {row_num} out of range (2-{ws.max_row}).", file=sys.stderr)
        sys.exit(1)

    # Append resolution to existing investigation
    existing = ws.cell(row=row_num, column=BUG_INVESTIGATION_COL).value or ""
    separator = "\n\n---\n**Resolution:** " if existing else "**Resolution:** "
    ws.cell(row=row_num, column=BUG_INVESTIGATION_COL, value=existing + separator + resolution).alignment = WRAP_ALIGNMENT
    ws.cell(row=row_num, column=BUG_STATUS_COL_INDEX, value=BUG_STATUS_FIXED).alignment = Alignment(
        horizontal="center", vertical="top"
    )
    wb.save(tracker)

    summary = ws.cell(row=row_num, column=BUG_SUMMARY_COL).value
    print(json.dumps({"status": "ok", "row": row_num, "summary": summary, "marked": BUG_STATUS_FIXED}))


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main() -> None:
    if len(sys.argv) < 2:
        print(__doc__, file=sys.stderr)
        sys.exit(1)

    cmd = sys.argv[1]

    if cmd == "init":
        if len(sys.argv) < 3:
            print("Usage: tracker.py init <dir> [handler ...]", file=sys.stderr)
            sys.exit(1)
        cmd_init(sys.argv[2], sys.argv[3:])

    elif cmd == "add":
        if len(sys.argv) < 7:
            print("Usage: tracker.py add <dir> <handler> <question> <internal_review> <status>", file=sys.stderr)
            sys.exit(1)
        cmd_add(sys.argv[2], sys.argv[3], sys.argv[4], sys.argv[5], sys.argv[6])

    elif cmd == "pending":
        handler = None
        if "--handler" in sys.argv:
            idx = sys.argv.index("--handler")
            if idx + 1 < len(sys.argv):
                handler = sys.argv[idx + 1]
        cmd_pending(sys.argv[2] if len(sys.argv) > 2 else ".", handler)

    elif cmd == "resolve":
        if len(sys.argv) < 5:
            print("Usage: tracker.py resolve <dir> <row_number> <answer>", file=sys.stderr)
            sys.exit(1)
        cmd_resolve(sys.argv[2], int(sys.argv[3]), sys.argv[4])

    elif cmd == "decide":
        if len(sys.argv) < 5:
            print("Usage: tracker.py decide <dir> <row_number> <decision+rationale>", file=sys.stderr)
            sys.exit(1)
        cmd_decide(sys.argv[2], int(sys.argv[3]), sys.argv[4])

    elif cmd == "add-handler":
        if len(sys.argv) < 4:
            print("Usage: tracker.py add-handler <dir> <handler>", file=sys.stderr)
            sys.exit(1)
        cmd_add_handler(sys.argv[2], sys.argv[3])

    elif cmd == "list-handlers":
        cmd_list_handlers(sys.argv[2] if len(sys.argv) > 2 else ".")

    elif cmd == "update-review":
        if len(sys.argv) < 6:
            print("Usage: tracker.py update-review <dir> <row> <internal_review> <status>", file=sys.stderr)
            sys.exit(1)
        cmd_update_review(sys.argv[2], int(sys.argv[3]), sys.argv[4], sys.argv[5])

    elif cmd == "doctor":
        if len(sys.argv) < 3:
            print("Usage: tracker.py doctor <dir>", file=sys.stderr)
            sys.exit(1)
        cmd_doctor(sys.argv[2])

    # ---- Bug commands ----
    elif cmd == "add-bug":
        if len(sys.argv) < 8:
            print("Usage: tracker.py add-bug <dir> <summary> <severity> <steps> <investigation> <status>",
                  file=sys.stderr)
            sys.exit(1)
        cmd_add_bug(sys.argv[2], sys.argv[3], sys.argv[4], sys.argv[5], sys.argv[6], sys.argv[7])

    elif cmd == "list-bugs":
        status_f = None
        severity_f = None
        show_all = "--all" in sys.argv
        if "--status" in sys.argv:
            idx = sys.argv.index("--status")
            if idx + 1 < len(sys.argv):
                status_f = sys.argv[idx + 1]
        if "--severity" in sys.argv:
            idx = sys.argv.index("--severity")
            if idx + 1 < len(sys.argv):
                severity_f = sys.argv[idx + 1]
        cmd_list_bugs(sys.argv[2] if len(sys.argv) > 2 else ".", status_f, severity_f, show_all)

    elif cmd == "update-bug":
        if len(sys.argv) < 6:
            print("Usage: tracker.py update-bug <dir> <row> <investigation> <status>", file=sys.stderr)
            sys.exit(1)
        cmd_update_bug(sys.argv[2], int(sys.argv[3]), sys.argv[4], sys.argv[5])

    elif cmd == "resolve-bug":
        if len(sys.argv) < 5:
            print("Usage: tracker.py resolve-bug <dir> <row> <resolution>", file=sys.stderr)
            sys.exit(1)
        cmd_resolve_bug(sys.argv[2], int(sys.argv[3]), sys.argv[4])

    else:
        print(f"Unknown command: {cmd}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
